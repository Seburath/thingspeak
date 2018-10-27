from openpyxl import load_workbook
import time

#acotacion dela funcion para obtener valores de excel
def cell(hoja, col, row):
    return hoja.cell(column=col, row=row).value


def w_cell(hoja, col, row, val):
    hoja.cell(column=col, row=row).value = val


def save(temperatura, radiacion, humedad):
    data = load_workbook(filename = 'data.xlsx')
    hoja = data["Hoja1"]

    y = hoja.max_row + 1

    w_cell(hoja, 1, y, time.asctime())
    w_cell(hoja, 2, y, temperatura)
    w_cell(hoja, 3, y, radiacion)
    w_cell(hoja, 4, y, humedad)

    print('registro numero: ' + str(hoja.max_row - 1) + ' añadido')
    data.save('data.xlsx')
